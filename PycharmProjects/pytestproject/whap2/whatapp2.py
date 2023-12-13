from selenium.webdriver.chrome.service import Service
from selenium import webdriver
options=webdriver.ChromeOptions
from selenium.webdriver.common.by import By
import time
# import openpyxl
# from datetime import datetime
# my_path='/Volumes/MAC - Data/excel/testdata2.xlsx'
# worksheet=openpyxl.load_workbook(my_path)
# current_sheet=worksheet.active
# row_data=current_sheet.max_row
# for i in range(2,current_sheet.max_row+1):
#     for j in range(1,current_sheet.max_column+1):
#         print(current_sheet.cell(row=i,column=j).value)



from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import openpyxl
from datetime import datetime

my_path = '/Volumes/MAC - Data/excel/testdata7.xlsx'
workbook = openpyxl.load_workbook(my_path)
current_sheet = workbook.active

# Get today's date without the time component
today_date = datetime.now().date()
print(today_date)
# Instantiate ChromeOptions
options = webdriver.ChromeOptions()
options.add_argument('user-data-dir=/Users/vaibhavlutade/Library/Caches/Google/Chrome/Default/Cache')

service = Service(executable_path='/Users/vaibhavlutade/PycharmProjects/pytestproject/whap2/chromedriver121')
# Navigate to the website
driver = webdriver.Chrome(service=service, options=options)

time.sleep(5)
credentials=[]

for i in range(2, current_sheet.max_row + 1):
    date_in_sheet = current_sheet.cell(row=i, column=1).value.date()
    name = current_sheet.cell(row=i, column=2).value
    credentials.append((date_in_sheet))
    print(credentials)


    date_in_sheet = datetime.strptime(str(date_in_sheet), "%Y-%d-%m").date()



    # Check if the date matches today's date
    if date_in_sheet == today_date:
        try:
            driver.get('https://web.whatsapp.com/')
            time.sleep(20)

            # time.sleep(20)  # Wait for 5 seconds
            driver.find_element(By.XPATH, "//*[@id='side']/div[1]/div/div[2]/div[2]/div/div[1]/p").send_keys(
                name)
            time.sleep(5)
            element1 = driver.find_elements(By.XPATH, "//div[@class='_21S-L']")
            time.sleep(5)
            # print(element1.tex)
            for i in element1:
                print(i.text)
                if i.text == name:
                    print(i)
                    i.click()
                    break
                # print(i.text)

            time.sleep(5)
            driver.find_element(By.XPATH,
                                "//*[@id='main']/footer/div[1]/div/span[2]/div/div[2]/div[1]/div/div[1]/p").send_keys(
                "hi rajesh")
            time.sleep(5)
            driver.find_element(By.XPATH, "//span[@data-icon='send']").click()
            time.sleep(5)
            driver.find_element(By.XPATH,
                                "//*[@id='main']/footer/div[1]/div/span[2]/div/div[1]/div[2]/div/div/div/span").click()
            time.sleep(5)
            sett = driver.find_element(By.XPATH, "//input[@accept='image/*,video/mp4,video/3gpp,video/quicktime']")
            time.sleep(5)
            path_data = "/Users/vaibhavlutade/Downloads/mahesh.jpeg"
            time.sleep(5)
            sett.send_keys(path_data)
            time.sleep(5)
            driver.find_element(By.XPATH,
                                "//*[@id='app']/div/div[2]/div[2]/div[2]/span/div/span/div/div/div[2]/div/div[2]/div[2]/div/div/span").click()
            time.sleep(5)


        except Exception as e:
            print("no matching date found")
        finally:

            driver.quit()
            workbook.close()


        # Your other Chrome WebDriver setup code here (if any)

        # Iterate through each cell and print its value

# Close the workbook when done
